from pathlib import Path
import tomllib
import argparse
from tqdm import tqdm
import json
import logging
from typing import Literal
import requests
import numpy as np
from numpy.linalg import inv
import openpyxl
import matplotlib.pyplot as plt
from skimage.segmentation import find_boundaries
import nrrd
from scipy.ndimage import affine_transform, label, binary_fill_holes, binary_closing

from read_scan import read_scan, read_bps

# in um
BRAIN_ORIGIN_CCFv3_COORD_POST_INF_ML = (5600, 5500, 5700)


def bincount_axes(x: np.ndarray, /, axis: int | tuple[int, ...], weights: np.ndarray = None) -> np.ndarray:
    """
    apply numpy.bincount on given axes of x
    """
    if weights is not None and x.shape != weights.shape:
        raise ValueError(f'x {x.shape} and weights {weights.shape} shape do not match')

    if isinstance(axis, int):
        axis = (axis,)

    # move flattened axes to the end
    dest = tuple(range(-len(axis), 0))
    x = np.moveaxis(x, axis, dest)

    batch_shape = x.shape[:-len(axis)]
    reduce_size = np.prod(x.shape[-len(axis):]).item()

    x = x.reshape(-1, reduce_size)

    k = x.max().item() + 1
    b = x.shape[0]

    offset = np.arange(b)[:, None] * k
    x = (x + offset).ravel()

    # apply same flatten to weights
    if weights is not None:
        weights = np.moveaxis(weights, axis, dest)
        weights = weights.reshape(-1, reduce_size)
        weights = weights.ravel()

    counts = np.bincount(x, weights=weights, minlength=b * k)

    counts = counts.reshape(*batch_shape, k)

    return counts


def find_only_file(root: Path, prefix: str, postfix: str) -> Path:
    pattern = f'{prefix}*{postfix}'
    result = list(root.glob(pattern))

    if len(result) == 0:
        raise ValueError(f'"{root}" does not contain {pattern}')
    if len(result) > 1:
        logging.warning(f'"{root}" contains {len(result)} of {pattern}, first one used')

    return result[0]


def download_annotation_volume(
        file_name: Path,
        ccf_version: Literal[2015, 2016, 2017, 2022] = 2022,
        resolution: Literal[10, 25, 50, 100] = 10
):
    url = (f'https://download.alleninstitute.org/informatics-archive/current-release/'
           f'mouse_ccf/annotation/ccf_{ccf_version}/annotation_{resolution}.nrrd')
    with requests.get(url, stream=True) as r:
        r.raise_for_status()
        with open(file_name, "wb") as f:
            for chunk in r.iter_content():
                f.write(chunk)


def download_allen_ontology(
        file_name: Path,
        structure_id: int
):
    url = f'https://api.brain-map.org/api/v2/structure_graph_download/{structure_id}.json'
    with requests.get(url) as r:
        r.raise_for_status()
        content = json.loads(r.content)
        if not content['success']:
            raise ValueError(f'ontology download failed')
        with open(file_name, "w") as f:
            json.dump(content['msg'][0], f)


def read_event_time(xlsx_path: Path):
    def group_tag(header: str) -> str:
        words = [w.lower() for w in header.split()]
        if words[0] != 'modified':
            raise ValueError(f'"{header}" is not valid')
        return words[1]

    wb = openpyxl.load_workbook(xlsx_path)
    if len(wb.sheetnames) == 1:
        ws = wb.worksheets[0]
    else:
        ws = wb['Processed Events']
    if ws.max_column != 2:
        raise ValueError(f'"{xlsx_path}" has {ws.max_column} columns')
    result = {}
    for i in range(2, ws.max_row + 1):
        event_type = ws.cell(row=i, column=1).value
        try:
            tag = group_tag(event_type)
        except ValueError:
            continue
        if tag in result:
            raise ValueError(f'more than one {tag} found in "{xlsx_path}"')
        values = ws.cell(row=i, column=2).value
        time = [int(s) for s in values.split()]
        time = np.array(time)
        result[tag] = time.reshape(-1, 2)
    return result


def pipeline(config: dict):
    ontology_path = Path(config['paths']['ontology'])
    if not ontology_path.is_file():
        download_allen_ontology(ontology_path, 1)  # adult mouse
    with open(ontology_path, 'r') as f:
        ontology = json.load(f)
    regions = {}
    queue = [ontology]
    while len(queue) > 0:
        item = queue.pop(0)
        regions[item['id']] = item  # reference in tree
        queue += item['children']

    annotation_path = Path(config['paths']['annotation'])
    if not annotation_path.is_file():
        download_annotation_volume(annotation_path)
    annotation_data, annotation_header = nrrd.read(annotation_path)

    annotation_transform = np.eye(4)
    annotation_transform[:3, :3] = annotation_header['space directions']
    annotation_transform[:3, 3] = annotation_header['space origin']

    # Iconeous "brain" coordinate is stereotaxic
    # direction: left, anterior, superior (dorsal)
    # origin: approx 4mm below Bregma
    # CCFv3 axis order: posterior, inferior, right
    # 'space' in header is incorrect
    # source: https://brain-map.org/support/documentation/api-allen-brain-connectivity-atlas
    brain_to_annotation = np.zeros((4, 4))
    brain_to_annotation[3, 3] = 1
    brain_to_annotation[0, 1] = -1
    brain_to_annotation[1, 2] = -1
    brain_to_annotation[2, 0] = -1
    brain_to_annotation[:3, 3] = BRAIN_ORIGIN_CCFv3_COORD_POST_INF_ML
    # magic number from Iconeous: 4mm (4000 um)
    brain_to_annotation[:3, :3] *= 4e3

    data_root = Path(config['paths']['data_root']).expanduser().resolve()
    for timepoint_dir in data_root.iterdir():
        if not timepoint_dir.is_dir(): continue
        timepoint = timepoint_dir.name
        event_time_dir = timepoint_dir / 'eventTime'
        for group_dir in event_time_dir.iterdir():
            if not group_dir.is_dir(): continue
            scan_dir = timepoint_dir / 'Scan' / group_dir.name.replace('eventTime_', 'Scan_')
            if not scan_dir.is_dir():
                raise ValueError(f'scan dir does not exist: "{scan_dir}"')
            group = group_dir.name.split('_')[2]
            # tqdm(list(group_dir.glob('*.xlsx')), desc=f"{timepoint},{group}")
            for event_time_path in group_dir.glob('[!~]*.xlsx'):
                # each trial
                parts = event_time_path.stem.split('_')
                subject = parts[0]
                prefix = '_'.join(parts[:-3])

                event_time = read_event_time(event_time_path)

                bps_path = find_only_file(scan_dir, prefix, '.source.bps')
                brain_to_lab = read_bps(bps_path)

                angio_scan_path = find_only_file(scan_dir, prefix, 'angio3D.source.scan')
                angio_scan = read_scan(angio_scan_path)

                fus_scan_path = find_only_file(scan_dir, prefix, 'fus3D.source.scan')
                fus_scan = read_scan(fus_scan_path)

                n_block_repeat = fus_scan.data.shape[2]
                if n_block_repeat != 1:
                    raise ValueError(f'block repeat number is {n_block_repeat}, we only handle 1 currently')
                data = fus_scan.data.squeeze(2)

                time = fus_scan.acquisition.time.squeeze(2)
                event_mask = {}
                for k, v in event_time.items():
                    event_mask[k] = ((time[..., None] >= v[:, 0]) & (time[..., None] <= v[:, 1])).any(axis=-1)

                threshold = np.percentile(data, 50, axis=(2, 3, 4), keepdims=True)
                mask = data > threshold
                # morphology process each 3d mask
                for i in range(data.shape[0]):
                    for j in range(data.shape[1]):
                        slice = mask[i, j]
                        labels, num = label(slice)
                        sizes = np.bincount(labels.ravel())
                        largest_label = np.argmax(sizes[1:]) + 1
                        body_mask = labels == largest_label
                        # ignore scan direction because it has much smaller size
                        for k in range(body_mask.shape[1]):
                            slice[:, k, :] = binary_closing(binary_fill_holes(body_mask[:, k, :]))

                data = np.log(data)
                per_body_norm = np.percentile(data, 99, axis=(2, 3, 4), keepdims=True)
                data /= per_body_norm

                voxels_to_annotation_index = (
                        inv(annotation_transform) @ brain_to_annotation @ inv(brain_to_lab) @
                        fus_scan.acquisition.probe_to_lab @ fus_scan.acquisition.voxels_to_probe
                )

                # shape: pose, x, y, z (no scan repeat because only depend on pose)
                voxel_annotations = np.empty(data.shape[1:], dtype=annotation_data.dtype)
                for i in range(data.shape[1]):
                    voxel_annotations[i] = affine_transform(
                        annotation_data,
                        matrix=voxels_to_annotation_index[i],
                        output_shape=data.shape[-3:],
                        order=0  # nearest neighbor
                    )

                body_3d_axes = (-3, -2, -1)
                # convert non-consecutive region ids to 0, 1, 2, ...
                ids, inverse = np.unique(voxel_annotations, return_inverse=True)
                inverse_b = np.broadcast_to(inverse, mask.shape)
                # shape: pose, id count
                region_voxel_count = bincount_axes(inverse, axis=body_3d_axes)
                # shape: repeat, pose, id count
                region_valid_voxel_count = bincount_axes(inverse_b, axis=body_3d_axes, weights=mask)
                region_valid_sum = bincount_axes(inverse_b, axis=body_3d_axes, weights=mask * data)

                # shape: scan repeat, pose, id count
                # contains nan if valid count is 0
                region_valid_ratio = region_valid_voxel_count / region_voxel_count[None, ...]
                valid_region_mask = region_valid_voxel_count > .8
                region_valid_mean = region_valid_sum / region_valid_voxel_count


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--config', type=Path, default=Path('config.toml'))
    args = parser.parse_args()

    if not args.config.is_file():
        raise FileNotFoundError(f'config file not found: {args.config}')
    with args.config.open('rb') as f:
        config = tomllib.load(f)

    pipeline(config)


if __name__ == "__main__":
    main()
