import openpyxl
import numpy as np
from pathlib import Path
import polars as pl

HEADER = 'Original Social Event'


def read_epochs(xlsx_path: str | Path) -> np.ndarray | None:
    wb = openpyxl.load_workbook(xlsx_path)
    if len(wb.sheetnames) == 1:
        ws = wb.worksheets[0]
    else:
        ws = wb['Processed Events']
    if ws.max_column != 2:
        raise ValueError(f'"{xlsx_path}" has {ws.max_column} columns')
    epochs = None
    for i in range(2, ws.max_row + 1):
        header = ws.cell(row=i, column=1).value
        if not header.startswith(HEADER):
            continue
        values = ws.cell(row=i, column=2).value
        time = [float(s) for s in values.split()]
        time = np.array(time)
        epochs = time.reshape(-1, 2)
        break
    return epochs


EVENT_NAME = 'event'
NON_EVENT_NAME = 'non-event'


def get_event_df(
        epochs: np.ndarray,
        total_time: float,
        hemodynamic_lag: float,
        max_event_n: int,
        min_event_time: float,
        max_event_time: float,
        post_event_exclusion_window: float,
        include_start_for_non_event: bool = False,
) -> tuple[pl.DataFrame, float]:
    def get_df(events: np.ndarray, type: str) -> pl.DataFrame:
        return pl.DataFrame({
            'onset': events[:, 0],
            'duration': events[:, 1],
            'trial_type': [type] * events.shape[0],
        })

    if epochs.ndim != 2 or epochs.shape[1] != 2:
        raise ValueError(f"epochs should have shape (n,2), got {epochs.shape}")

    epochs += hemodynamic_lag
    event_n = epochs.shape[0]
    non_event_epochs = np.empty((event_n + 1, 2), dtype=epochs.dtype)
    non_event_epochs[0, 0] = .0
    non_event_epochs[-1, 1] = total_time
    non_event_epochs[1:, 0] = epochs[:, 1] + post_event_exclusion_window
    non_event_epochs[:-1, 1] = epochs[:, 0]
    if max_event_n >= event_n:
        max_time = total_time
    else:
        max_time = epochs[max_event_n, 0]
        epochs = epochs[:max_event_n, :]
        non_event_epochs = non_event_epochs[:max_event_n + 1, :]
    epochs[:, 1] -= epochs[:, 0]
    epochs = epochs[epochs[:, 1] >= min_event_time]
    epochs[epochs[:, 1] > max_event_time, 1] = max_event_time
    event_df = get_df(epochs, EVENT_NAME)

    non_event_epochs = non_event_epochs[int(include_start_for_non_event):, :]
    non_event_epochs[:, 1] -= non_event_epochs[:, 0]
    non_event_df = get_df(non_event_epochs, NON_EVENT_NAME)

    return pl.concat([event_df, non_event_df]), max_time
